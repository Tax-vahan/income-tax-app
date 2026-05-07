"""
Excel output service — completely isolated from all network/Selenium logic.

Produces a 4-sheet workbook:
  Sheet 1 — Challan Details   (one row per challan, sortable/filterable)
  Sheet 2 — Section Summary   (group-by section code + nature of payment)
  Sheet 3 — Month Summary     (group-by month-year)
  Sheet 4 — Run Info          (execution metadata)
"""

import os
import logging
from collections import defaultdict
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

log = logging.getLogger("TDS")


# ── Design tokens ─────────────────────────────────────────────────────────────

_C = {
    "hdr_bg":   "1F4E79",
    "hdr_fg":   "FFFFFF",
    "title_bg": "002060",
    "alt":      "EBF3FB",
    "tot_bg":   "FCE4D6",
    "border":   "B8CCE4",
}

_TH  = Side(style="thin",   color=_C["border"])
_MED = Side(style="medium", color=_C["hdr_bg"])
_B   = Border(left=_TH, right=_TH, top=_TH, bottom=_TH)
_BB  = Border(left=_TH, right=_TH, top=_TH, bottom=_MED)


# ── Column spec — (header, field_key, width, number_format) ──────────────────

COLUMNS = [
    ("Sr. No.",           "__sr__",        7,  None),
    ("BSR Code",          "bsrCode",      12,  None),
    ("Challan No.",       "challanNum",   13,  None),
    ("Payment Date",      "paymentDt",    16,  None),
    ("Financial Year",    "financialYear",14,  None),
    ("Asst. Year",        "assessmentYear",12, None),
    ("Section Code",      "secCd",        11,  None),
    ("Minor Head",        "minorHead",    11,  None),
    ("CIN",               "cin",          24,  None),
    ("Bank",              "bankCode",      8,  None),
    ("Basic Tax (Rs.)",   "basicTax",     16,  "#,##0"),
    ("Surcharge (Rs.)",   "surCharge",    14,  "#,##0"),
    ("Edu Cess (Rs.)",    "eduCess",      13,  "#,##0"),
    ("Interest (Rs.)",    "interest",     13,  "#,##0"),
    ("Penalty (Rs.)",     "penalty",      13,  "#,##0"),
    ("Others (Rs.)",      "others",       12,  "#,##0"),
    ("Total Amt (Rs.)",   "totalAmt",     18,  "#,##0"),
]

_AMT_FIELDS = {
    "basicTax", "surCharge", "eduCess",
    "interest", "penalty", "others", "totalAmt",
}


# ── Low-level cell writers ────────────────────────────────────────────────────

def _hcell(ws, row, col, val):
    c = ws.cell(row, col, val)
    c.font      = Font(name="Arial", bold=True, color=_C["hdr_fg"], size=9)
    c.fill      = PatternFill("solid", fgColor=_C["hdr_bg"])
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = _B


def _dcell(ws, row, col, val, fmt=None, alt=False):
    c = ws.cell(row, col, val)
    c.font      = Font(name="Arial", size=9)
    c.alignment = Alignment(vertical="center")
    c.border    = _B
    if alt:
        c.fill = PatternFill("solid", fgColor=_C["alt"])
    if fmt:
        c.number_format = fmt


def _tcell(ws, row, col, val, fmt=None):
    c = ws.cell(row, col, val)
    c.font      = Font(name="Arial", bold=True, size=10)
    c.fill      = PatternFill("solid", fgColor=_C["tot_bg"])
    c.alignment = Alignment(horizontal="right", vertical="center")
    c.border    = _BB
    if fmt:
        c.number_format = fmt


def _title(ws, text, ncols, row=1, height=28):
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row,   end_column=ncols)
    c = ws.cell(row, 1, text)
    c.font      = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    c.fill      = PatternFill("solid", fgColor=_C["title_bg"])
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = height


# ── Sheet builders ────────────────────────────────────────────────────────────

def _detail_sheet(wb, challans, cfg):
    ws = wb.active
    ws.title        = "Challan Details"
    ws.freeze_panes = "D3"
    nc = len(COLUMNS)

    _title(
        ws,
        f"TDS Challan Details  |  TAN: {cfg['TAN']}  |  "
        f"Period: {cfg['FROM_DATE']} to {cfg['TO_DATE']}  |  "
        f"Generated: {datetime.now().strftime('%d-%b-%Y %H:%M')}",
        nc,
    )
    ws.row_dimensions[2].height = 36

    for ci, (hdr, _, w, _) in enumerate(COLUMNS, 1):
        _hcell(ws, 2, ci, hdr)
        ws.column_dimensions[get_column_letter(ci)].width = w

    # Fields that must always be stored as text to preserve leading zeros
    _TEXT_FIELDS = {"bsrCode", "challanNum", "cin"}

    for ri, ch in enumerate(challans, 3):
        ws.row_dimensions[ri].height = 16
        alt = (ri % 2 == 0)
        for ci, (_, field, _, fmt) in enumerate(COLUMNS, 1):
            if field == "__sr__":
                val = ri - 2
            else:
                val = ch.get(field) or ""
                # Coerce identifier fields to str so Excel never interprets them
                # as numbers (which would silently drop leading zeros)
                if field in _TEXT_FIELDS and val != "":
                    val = str(val)
            _dcell(ws, ri, ci, val, fmt=fmt, alt=alt)

    tr = 3 + len(challans)
    ws.row_dimensions[tr].height = 20
    c = ws.cell(tr, 1, "TOTAL")
    c.font = Font(name="Arial", bold=True, size=10)
    c.fill = PatternFill("solid", fgColor=_C["tot_bg"])

    for ci, (_, field, _, _fmt) in enumerate(COLUMNS, 1):
        if field in _AMT_FIELDS:
            cl = get_column_letter(ci)
            _tcell(ws, tr, ci, f"=SUM({cl}3:{cl}{tr-1})", "#,##0")
        elif ci > 1:
            _tcell(ws, tr, ci, "")

    ws.auto_filter.ref = f"A2:{get_column_letter(nc)}{tr-1}"


def _section_sheet(wb, challans, grand):
    ws  = wb.create_sheet("Section Summary")
    sh  = [("Section", 14), ("Nature of Payment", 28),
           ("Count", 10), ("Total Amount (Rs.)", 20), ("% of Total", 14)]
    _title(ws, "Section-wise Summary", len(sh))
    ws.row_dimensions[2].height = 28
    for ci, (hdr, w) in enumerate(sh, 1):
        _hcell(ws, 2, ci, hdr)
        ws.column_dimensions[get_column_letter(ci)].width = w

    agg = defaultdict(lambda: {"desc": "", "count": 0, "total": 0})
    for ch in challans:
        k = ch.get("secCd") or "N/A"
        agg[k]["desc"]   = ch.get("natrPymntDesc") or ""
        agg[k]["count"] += 1
        agg[k]["total"] += ch.get("totalAmt", 0) or 0

    r = 3
    for k, v in sorted(agg.items()):
        alt = (r % 2 == 0)
        pct = round(v["total"] / grand * 100, 2) if grand else 0.0
        _dcell(ws, r, 1, k,         alt=alt)
        _dcell(ws, r, 2, v["desc"], alt=alt)
        _dcell(ws, r, 3, v["count"],alt=alt)
        _dcell(ws, r, 4, v["total"],fmt="#,##0", alt=alt)
        _dcell(ws, r, 5, pct,       fmt="0.00",  alt=alt)
        r += 1

    _tcell(ws, r, 1, "GRAND TOTAL")
    _tcell(ws, r, 2, "")
    _tcell(ws, r, 3, f"=SUM(C3:C{r-1})")
    _tcell(ws, r, 4, f"=SUM(D3:D{r-1})", "#,##0")
    _tcell(ws, r, 5, "100.00")


def _month_sheet(wb, challans):
    ws  = wb.create_sheet("Month Summary")
    sh  = [("Month-Year", 16), ("Count", 10), ("Total Amount (Rs.)", 20)]
    _title(ws, "Month-wise Summary", len(sh))
    ws.row_dimensions[2].height = 28
    for ci, (hdr, w) in enumerate(sh, 1):
        _hcell(ws, 2, ci, hdr)
        ws.column_dimensions[get_column_letter(ci)].width = w

    agg = defaultdict(lambda: {"count": 0, "total": 0, "_dt": datetime.min})
    for ch in challans:
        # Prefer the pre-computed paymentMonth field ("Apr 2026"); fall back to paymentDt parsing
        key = ch.get("paymentMonth") or ""
        dt  = None
        if not key:
            pt = ch.get("paymentDt") or ""
            for fmt, slen in (("%d/%m/%Y", 10), ("%d-%b-%Y", 11), ("%d-%m-%Y", 10)):
                try:
                    dt  = datetime.strptime(pt[:slen].strip(), fmt)
                    key = dt.strftime("%b %Y")
                    break
                except Exception:
                    continue
        if not key:
            key = "Unknown"
        if dt is None and key != "Unknown":
            try:
                dt = datetime.strptime(key, "%b %Y")
            except Exception:
                dt = datetime.min
        agg[key]["count"] += 1
        agg[key]["total"] += ch.get("totalAmt", 0) or 0
        if dt and dt > agg[key]["_dt"]:
            agg[key]["_dt"] = dt

    r = 3
    for key in sorted(agg, key=lambda x: agg[x]["_dt"]):
        alt = (r % 2 == 0)
        _dcell(ws, r, 1, key,               alt=alt)
        _dcell(ws, r, 2, agg[key]["count"], alt=alt)
        _dcell(ws, r, 3, agg[key]["total"], fmt="#,##0", alt=alt)
        r += 1

    _tcell(ws, r, 1, "TOTAL")
    _tcell(ws, r, 2, f"=SUM(B3:B{r-1})")
    _tcell(ws, r, 3, f"=SUM(C3:C{r-1})", "#,##0")


def _run_info_sheet(wb, challans, cfg, grand, stats):
    ws = wb.create_sheet("Run Info")
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 55
    _title(ws, "Run Information", 2)

    rows = [
        ("Parameter",             "Value"),
        ("TAN",                   cfg["TAN"]),
        ("From Date",             cfg["FROM_DATE"]),
        ("To Date",               cfg["TO_DATE"]),
        ("Total Challans",        len(challans)),
        ("Grand Total (Rs.)",     f"{grand:,.0f}"),
        ("Detail Fetch Mode",     cfg.get("DETAIL_MODE", "AUTO")),
        ("Detail Success Rate",   f"{stats.get('detail_success_rate', 'N/A')}%"),
        ("Execution Time (s)",    stats.get("duration", "N/A")),
        ("Generated On",          datetime.now().strftime("%d-%b-%Y %H:%M:%S")),
        ("Portal",                "https://eportal.incometax.gov.in"),
        ("Script Version",        "v6.0 — modular CDP + requests.Session"),
    ]
    for ri, (k, v) in enumerate(rows, 2):
        c1 = ws.cell(ri, 1, k)
        c2 = ws.cell(ri, 2, str(v))
        if ri == 2:
            for c in (c1, c2):
                c.font  = Font(name="Arial", bold=True, color="FFFFFF", size=9)
                c.fill  = PatternFill("solid", fgColor=_C["hdr_bg"])
        else:
            c1.font = Font(name="Arial", bold=True, size=9)
            c2.font = Font(name="Arial", size=9)
            if ri % 2 == 0:
                for c in (c1, c2):
                    c.fill = PatternFill("solid", fgColor=_C["alt"])
        for c in (c1, c2):
            c.border    = _B
            c.alignment = Alignment(vertical="center")


# ── Public entry point ────────────────────────────────────────────────────────

def write_excel(
    challans: list,
    cfg:      dict,
    stats:    dict | None = None,
) -> tuple[str, int]:
    """
    Write challans to a formatted .xlsx workbook.

    Returns (output_path, grand_total_int).
    """
    if stats is None:
        stats = {}

    grand = sum(ch.get("totalAmt", 0) or 0 for ch in challans)

    wb = Workbook()
    _detail_sheet(wb, challans, cfg)
    _section_sheet(wb, challans, grand)
    _month_sheet(wb, challans)
    _run_info_sheet(wb, challans, cfg, grand, stats)

    out = cfg.get("OUTPUT_PATH") or os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        "..", "..",             # back to tds_api/
        cfg["OUTPUT_FILE"],
    )
    out = os.path.normpath(out)
    wb.save(out)
    log.info(
        "Excel saved → %s  (%d challans, grand total %s)",
        out, len(challans), f"{grand:,.0f}",
    )
    return out, grand
